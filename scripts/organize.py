#!/usr/bin/env python3
"""
文档归类整理工具 v6
只做两件事：
1. 商务信息Excel按客户拆分 → 客户档案/{客户}/对应子目录（覆盖）
2. 运维工单Excel增量拆分合并 → 客户档案/{客户}/运维工单/（去重合并）
"""

import os
import json
import hashlib
import shutil
import argparse
import logging
from pathlib import Path
from datetime import datetime
import pandas as pd
import openpyxl
import xlrd

# === 配置 ===
CONFIG_FILE = Path(__file__).parent.parent / 'config.json'
CLIENT_DATA_ROOT = Path('/Users/limingheng/AI/client-data')
SOURCE_DIR = CLIENT_DATA_ROOT / 'raw'
TARGET_DIR = CLIENT_DATA_ROOT / '客户档案'
MAPPING_FILE = TARGET_DIR / '_mapping.json'
LOG_DIR = Path(__file__).parent.parent / 'logs'


def load_config():
    if CONFIG_FILE.exists():
        with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}


def get_path(key, default):
    cfg = load_config()
    val = cfg.get(key, default)
    return Path(val) if val else Path(default)


# === 日志 ===
def setup_logging():
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    log_file = LOG_DIR / f'organize_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log'
    logger = logging.getLogger('doc-organizer')
    logger.setLevel(logging.INFO)
    logger.handlers = []
    fh = logging.FileHandler(log_file, encoding='utf-8')
    fh.setFormatter(logging.Formatter('[%(asctime)s] %(message)s', '%H:%M:%S'))
    ch = logging.StreamHandler()
    ch.setFormatter(logging.Formatter('[%(asctime)s] %(message)s', '%H:%M:%S'))
    logger.addHandler(fh)
    logger.addHandler(ch)
    return logger, log_file


# === 映射表 ===
def load_mapping():
    if MAPPING_FILE.exists():
        with open(MAPPING_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}


def get_short_name(full_name, mapping):
    """全称 → 简称（反向查mapping）"""
    for short, full in mapping.items():
        if full == full_name:
            return short
    return None


def get_file_hash(filepath):
    h = hashlib.md5()
    with open(filepath, 'rb') as f:
        for chunk in iter(lambda: f.read(8192), b''):
            h.update(chunk)
    return h.hexdigest()


# === 任务1：商务信息Excel拆分（覆盖）===
BIZ_FILES = [
    {
        'src_glob': '客户主数据/*.xlsx',
        'dst_subdir': '基础数据',
        'dst_filename': '客户主数据.xlsx',
        'client_col': '真实服务对象',
    },
    {
        'src_glob': '订阅台账/*.xlsx',
        'dst_subdir': '订阅合同行',
        'dst_filename': '订阅台账.xlsx',
        'client_col': '真实服务对象',
    },
    {
        'src_glob': '固定金额台账/*.xlsx',
        'dst_subdir': '实施合同行',
        'dst_filename': '固定金额台账.xlsx',
        'client_col': '最终服务对象',
    },
    {
        'src_glob': '人天框架台账/*.xlsx',
        'dst_subdir': '实施合同行',
        'dst_filename': '人天框架台账.xlsx',
        'client_col': '最终服务对象',
    },
    {
        'src_glob': '项目收款进度查询/*',
        'dst_subdir': '订阅合同收款情况',
        'dst_filename': '项目收款进度查询.xlsx',
        'client_col': '最终服务对象名称',
    },
]


def split_excel_openpyxl(src_file, mapping, spec, logger):
    """用openpyxl读取xlsx，按客户列拆分，一次性写入"""
    wb = openpyxl.load_workbook(src_file)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_row = [c.value for c in ws[1]]
        client_col_idx = None
        for i, col in enumerate(header_row):
            if col and spec['client_col'] in str(col):
                client_col_idx = i
                break
        if client_col_idx is None:
            continue

        # 先把所有行按客户聚合
        client_rows = {}  # {short_name: [row_data, ...]}
        for row_idx in range(2, ws.max_row + 1):
            cell = ws[row_idx][client_col_idx]
            if not cell.value:
                continue
            full_name = str(cell.value).strip()
            short_name = get_short_name(full_name, mapping)
            if not short_name:
                continue
            row_data = [c.value if c.value is not None else '' for c in ws[row_idx]]
            client_rows.setdefault(short_name, []).append(row_data)

        # 逐客户写入
        for short_name, rows in client_rows.items():
            dst_file = TARGET_DIR / short_name / spec['dst_subdir'] / spec['dst_filename']
            dst_file.parent.mkdir(parents=True, exist_ok=True)
            new_df = pd.DataFrame(rows, columns=header_row)
            if dst_file.exists():
                try:
                    existing = pd.read_excel(dst_file)
                    df = pd.concat([existing, new_df], ignore_index=True)
                except Exception:
                    df = new_df
            else:
                df = new_df
            df.to_excel(dst_file, index=False)
            logger.info(f'  更新 {short_name}/{spec["dst_subdir"]}/{spec["dst_filename"]}')


def split_excel_xlrd(src_file, mapping, spec, logger):
    """用xlrd读取xls，按客户列拆分，一次性写入"""
    wb = xlrd.open_workbook(src_file)
    ws = wb.sheet_by_index(0)
    header_row = [ws.cell_value(0, col) for col in range(ws.ncols)]
    client_col_idx = None
    for i, col in enumerate(header_row):
        if col and spec['client_col'] in str(col):
            client_col_idx = i
            break
    if client_col_idx is None:
        return

    client_rows = {}
    for row_idx in range(1, ws.nrows):
        full_name = str(ws.cell_value(row_idx, client_col_idx)).strip()
        if not full_name:
            continue
        short_name = get_short_name(full_name, mapping)
        if not short_name:
            continue
        row_data = [ws.cell_value(row_idx, col) for col in range(ws.ncols)]
        client_rows.setdefault(short_name, []).append(row_data)

    for short_name, rows in client_rows.items():
        dst_file = TARGET_DIR / short_name / spec['dst_subdir'] / spec['dst_filename']
        dst_file.parent.mkdir(parents=True, exist_ok=True)
        new_df = pd.DataFrame(rows, columns=header_row)
        if dst_file.exists():
            try:
                existing = pd.read_excel(dst_file)
                df = pd.concat([existing, new_df], ignore_index=True)
            except Exception:
                df = new_df
        else:
            df = new_df
        df.to_excel(dst_file, index=False)
        logger.info(f'  更新 {short_name}/{spec["dst_subdir"]}/{spec["dst_filename"]}')
def organize_business_summary(logger, mapping):
    """商务信息Excel拆分"""
    logger.info('=== 商务信息拆分 ===')
    biz_dir = SOURCE_DIR / '商务信息'
    if not biz_dir.exists():
        logger.info(f'  目录不存在: {biz_dir}')
        return

    for spec in BIZ_FILES:
        files = list(biz_dir.glob(spec['src_glob']))
        if not files:
            logger.info(f'  无文件: {spec["src_glob"]}')
            continue
        for f in files:
            if f.name.startswith('test') or f.name.startswith('~'):
                continue
            logger.info(f'  处理: {f.name}')
            try:
                if f.suffix == '.xls':
                    split_excel_xlrd(f, mapping, spec, logger)
                else:
                    split_excel_openpyxl(f, mapping, spec, logger)
            except Exception as e:
                logger.error(f'  错误 {f.name}: {e}')


# === 任务2：运维工单增量拆分合并 ===
def organize_work_orders(logger, mapping):
    """运维工单增量拆分（按编号去重后合并）"""
    logger.info('=== 运维工单拆分 ===')
    op_dir = SOURCE_DIR / '运维工单'
    if not op_dir.exists():
        logger.info(f'  目录不存在: {op_dir}')
        return

    for f in op_dir.glob('*.xlsx'):
        logger.info(f'  处理: {f.name}')
        try:
            xl = pd.ExcelFile(f)
            df = None
            for sheet in xl.sheet_names:
                try:
                    _df = xl.parse(sheet)
                    for col in _df.columns:
                        if '客户' in str(col) and '名称' in str(col) and '期望' not in str(col) and '区域' not in str(col) and '经理' not in str(col):
                            df = _df
                            break
                    if df is not None:
                        break
                except Exception:
                    continue
            if df is None or df.empty:
                logger.info(f'  无客户名称列: {f.name}')
                continue

            # 找客户列
            client_col = None
            for col in df.columns:
                if '客户' in str(col) and '名称' in str(col) and '期望' not in str(col) and '区域' not in str(col) and '经理' not in str(col):
                    client_col = col
                    break

            if not client_col:
                continue

            for full_name in df[client_col].dropna().unique():
                full_name = str(full_name).strip()
                short_name = get_short_name(full_name, mapping)
                if not short_name:
                    continue

                client_df = df[df[client_col] == full_name]
                dst_file = TARGET_DIR / short_name / '运维工单' / f.name
                dst_file.parent.mkdir(parents=True, exist_ok=True)

                if dst_file.exists():
                    try:
                        existing = pd.read_excel(dst_file)
                        if '编号' in existing.columns and '编号' in client_df.columns:
                            combined = pd.concat([existing, client_df]).drop_duplicates(subset=['编号'], keep='last')
                        else:
                            combined = pd.concat([existing, client_df])
                        combined.to_excel(dst_file, index=False)
                    except Exception:
                        client_df.to_excel(dst_file, index=False)
                else:
                    client_df.to_excel(dst_file, index=False)

            logger.info(f'  完成: {f.name}')
        except Exception as e:
            logger.error(f'  错误 {f.name}: {e}')


def main():
    parser = argparse.ArgumentParser(description='文档归类整理工具 v6')
    parser.add_argument('--dry-run', action='store_true', help='预览模式')
    args = parser.parse_args()

    logger, log_file = setup_logging()

    if args.dry_run:
        logger.info('【预览模式】不做任何实际写入操作')
        logger.info(f'SOURCE_DIR: {SOURCE_DIR}')
        logger.info(f'TARGET_DIR: {TARGET_DIR}')
        logger.info(f'MAPPING_FILE: {MAPPING_FILE}')
        return

    mapping = load_mapping()
    if not mapping:
        logger.error('映射表为空，请先运行生成映射表')
        return

    logger.info(f'映射表加载: {len(mapping)} 个客户')

    logger.info('')
    organize_business_summary(logger, mapping)

    logger.info('')
    organize_work_orders(logger, mapping)

    logger.info('')
    logger.info('=== 完成 ===')
    logger.info(f'日志: {log_file}')


if __name__ == '__main__':
    main()
